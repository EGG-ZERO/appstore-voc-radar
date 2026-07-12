"""把VOC监测数据导出为带格式的Excel周报（业务侧交付形态）。

三个工作表：
- 概览：KPI块、各App近7日健康度、投诉类别Top10（数据条）、类别柱状图
- 评论明细：负面评论优先展示，冻结首行+自动筛选+隔行着色
- 竞品x类别：投诉类别×App交叉计数

用法：python src/export_excel.py（先跑analyze.py）
产出：reports/竞品VOC周报.xlsx
"""
from __future__ import annotations

import csv
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
STATS_JSON = PROJECT_ROOT / "data" / "daily_stats.json"
REVIEWS_CSV = PROJECT_ROOT / "data" / "reviews.csv"
OUT = PROJECT_ROOT / "reports" / "竞品VOC周报.xlsx"

FONT = "Microsoft YaHei"
DARK = "1F4E79"
LIGHT_FILL = PatternFill("solid", fgColor="EDF2F9")
HEADER_FILL = PatternFill("solid", fgColor=DARK)
HEADER_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT, size=16, bold=True, color=DARK)
KPI_FONT = Font(name=FONT, size=22, bold=True, color=DARK)
LABEL_FONT = Font(name=FONT, size=9, color="6B7688")
BODY_FONT = Font(name=FONT, size=10)
NEG_FONT = Font(name=FONT, size=10, bold=True, color="B94040")
THIN = Border(*[Side(style="thin", color="D8DEE8")] * 4)


def style_header(ws, row: int, col_start: int, col_end: int) -> None:
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
        cell.alignment = Alignment(horizontal="center", vertical="center")


def recent_by_app(stats: dict, days: int = 28) -> list[dict]:
    dates = sorted(stats["daily"].keys())
    if not dates:
        return []
    cutoff = (datetime.strptime(dates[-1], "%Y-%m-%d")
              - timedelta(days=days - 1)).strftime("%Y-%m-%d")
    rows = []
    for app_id, name in stats["apps"].items():
        n = neg = rsum = 0
        for d in dates:
            if d >= cutoff and app_id in stats["daily"][d]:
                day = stats["daily"][d][app_id]
                n += day["n"]; neg += day["neg"]; rsum += day["rating_sum"]
        if n:
            rows.append({"app": name, "n": n, "neg_rate": neg / n, "avg": rsum / n})
    return sorted(rows, key=lambda r: -r["n"])


def sheet_overview(wb: Workbook, stats: dict) -> None:
    ws = wb.active
    ws.title = "概览"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "竞品VOC周报 · App Store中国区评论监测"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = ("咖啡茶饮类目5个App · 负面=1-2星 · 评论样本有幸存者偏差（趋势与对比有效）· "
                f"生成于 {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC")
    ws["B3"].font = LABEL_FONT

    sig = sum(1 for t in stats["version_tests"] if t["significant"])
    kpis = [("累计评论", stats["total_reviews"]),
            ("本次新增", stats["new_today"]),
            ("活跃预警", len(stats["alerts"])),
            ("版本检验显著", f"{sig}/{len(stats['version_tests'])}")]
    for i, (label, value) in enumerate(kpis):
        col = 2 + i * 2
        ws.cell(row=5, column=col, value=label).font = LABEL_FONT
        ws.cell(row=6, column=col, value=value).font = KPI_FONT

    ws["B9"] = "各App近28天健康度"
    ws["B9"].font = Font(name=FONT, size=12, bold=True, color=DARK)
    for i, h in enumerate(["App", "评论数", "负面率", "平均分"]):
        ws.cell(row=10, column=2 + i, value=h)
    style_header(ws, 10, 2, 5)
    r7 = recent_by_app(stats)
    for r, row in enumerate(r7, start=11):
        ws.cell(row=r, column=2, value=row["app"]).font = BODY_FONT
        ws.cell(row=r, column=3, value=row["n"]).font = BODY_FONT
        c = ws.cell(row=r, column=4, value=row["neg_rate"])
        c.number_format = "0.0%"
        c.font = NEG_FONT if row["neg_rate"] >= 0.5 else BODY_FONT
        a = ws.cell(row=r, column=5, value=round(row["avg"], 2))
        a.font = BODY_FONT
        for col in (2, 3, 4, 5):
            ws.cell(row=r, column=col).border = THIN

    ws["G9"] = "负面投诉类别Top10（全部App合计）"
    ws["G9"].font = Font(name=FONT, size=12, bold=True, color=DARK)
    totals: dict[str, int] = {}
    for cats in stats["cats_by_app"].values():
        for cat, cnt in cats.items():
            totals[cat] = totals.get(cat, 0) + cnt
    top = sorted(totals.items(), key=lambda kv: -kv[1])[:10]
    for i, h in enumerate(["类别", "提及数"]):
        ws.cell(row=10, column=7 + i, value=h)
    style_header(ws, 10, 7, 8)
    for r, (cat, cnt) in enumerate(top, start=11):
        ws.cell(row=r, column=7, value=cat).font = BODY_FONT
        ws.cell(row=r, column=8, value=cnt).font = BODY_FONT
        for col in (7, 8):
            ws.cell(row=r, column=col).border = THIN
    last = 10 + len(top)
    if top:
        ws.conditional_formatting.add(
            f"H11:H{last}",
            DataBarRule(start_type="num", start_value=0, end_type="max",
                        color=DARK, showValue=True))
        chart = BarChart()
        chart.type = "bar"
        chart.title = "投诉类别Top10（负面评论提及数）"
        chart.height, chart.width = 8.5, 14
        chart.legend = None
        data = Reference(ws, min_col=8, min_row=10, max_row=last)
        cats_ref = Reference(ws, min_col=7, min_row=11, max_row=last)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "B18")

    for col, width in {"A": 2, "B": 13, "C": 9, "D": 9, "E": 9, "F": 3,
                       "G": 13, "H": 9}.items():
        ws.column_dimensions[col].width = width


def sheet_details(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("评论明细")
    headers = ["日期", "App", "评分", "版本", "投诉类别", "标题", "内容"]
    keys = ["date", "app_name", "rating", "version", "cats", "title", "content"]
    ws.append(headers)
    style_header(ws, 1, 1, len(headers))
    # 负面在前、日期新在前
    ordered = sorted(rows, key=lambda r: (int(r["rating"]) > 2, r["date"]), reverse=False)
    ordered = sorted(ordered, key=lambda r: r["date"], reverse=True)
    ordered = sorted(ordered, key=lambda r: int(r["rating"]))
    for i, row in enumerate(ordered):
        vals = [row[k].replace("|", "、") if k == "cats" else
                (row[k][:120] if k == "content" else row[k]) for k in keys]
        ws.append(vals)
        r = i + 2
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = THIN
            if i % 2 == 1:
                cell.fill = LIGHT_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(ordered) + 1}"
    for col, width in {"A": 11, "B": 12, "C": 6, "D": 10, "E": 18,
                       "F": 28, "G": 60}.items():
        ws.column_dimensions[col].width = width


def sheet_crosstab(wb: Workbook, stats: dict) -> None:
    ws = wb.create_sheet("竞品x类别")
    apps = sorted(stats["cats_by_app"].keys())
    cats = sorted({c for m in stats["cats_by_app"].values() for c in m},
                  key=lambda c: -sum(m.get(c, 0) for m in stats["cats_by_app"].values()))
    ws.append(["投诉类别"] + apps + ["合计"])
    style_header(ws, 1, 1, len(apps) + 2)
    for i, cat in enumerate(cats):
        vals = [stats["cats_by_app"][a].get(cat, 0) for a in apps]
        ws.append([cat] + vals + [sum(vals)])
        for c in range(1, len(apps) + 3):
            cell = ws.cell(row=i + 2, column=c)
            cell.font = BODY_FONT
            cell.border = THIN
    ws.column_dimensions["A"].width = 14
    for c in range(2, len(apps) + 3):
        ws.column_dimensions[get_column_letter(c)].width = 12


def main() -> None:
    stats = json.loads(STATS_JSON.read_text(encoding="utf-8"))
    with open(REVIEWS_CSV, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))

    wb = Workbook()
    sheet_overview(wb, stats)
    sheet_details(wb, rows)
    sheet_crosstab(wb, stats)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Excel周报已生成: {OUT} （{len(rows)}条评论明细）")


if __name__ == "__main__":
    main()
